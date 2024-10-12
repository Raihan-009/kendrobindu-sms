from fastapi import FastAPI, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract, func
from . import models, schemas, database
from typing import List
from datetime import date, timedelta
import random
import string
from fastapi.encoders import jsonable_encoder
import logging
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os
from openpyxl.chart import LineChart, Reference

app = FastAPI()

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create tables
database.create_tables()

def generate_unique_id(hsc_batch):
    letters = string.ascii_uppercase + string.digits
    unique_part = ''.join(random.choice(letters) for _ in range(6))
    return f"{unique_part}-{hsc_batch}"

@app.post("/reset-database")
def reset_database():
    database.reset_database()
    return {"message": "Database reset successfully"}

@app.post("/students/", response_model=schemas.Student)
def create_student(student: schemas.StudentCreate, db: Session = Depends(database.get_db)):
    unique_id = generate_unique_id(student.hsc_batch)
    db_student = models.Student(id=unique_id, **student.dict())
    db.add(db_student)
    db.commit()
    db.refresh(db_student)
    return db_student

@app.get("/students/", response_model=List[schemas.Student])
def get_students(skip: int = 0, limit: int = 100, db: Session = Depends(database.get_db)):
    students = db.query(models.Student).offset(skip).limit(limit).all()
    return students

@app.get("/students/{student_id}", response_model=schemas.Student)
def get_student(student_id: str, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")
    return student

@app.get("/students/batch/{kb_batch}", response_model=List[schemas.Student])
def get_students_by_batch(kb_batch: str, db: Session = Depends(database.get_db)):
    students = db.query(models.Student).filter(models.Student.kb_batch == kb_batch).all()
    if not students:
        raise HTTPException(status_code=404, detail="No students found for this batch")
    return students

@app.post("/attendance/", response_model=schemas.Attendance)
def create_attendance(attendance: schemas.AttendanceCreate, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == attendance.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    existing_attendance = db.query(models.Attendance).filter(
        models.Attendance.student_id == attendance.student_id,
        models.Attendance.date == attendance.date
    ).first()
    
    if existing_attendance:
        existing_attendance.present = attendance.present
        db.commit()
        db.refresh(existing_attendance)
        return existing_attendance
    else:
        db_attendance = models.Attendance(**attendance.dict())
        db.add(db_attendance)
        db.commit()
        db.refresh(db_attendance)
        return db_attendance

@app.post("/payments/", response_model=schemas.PaymentHistory)
def create_payment(payment: schemas.PaymentHistoryCreate, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == payment.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    due = payment.payment - payment.paid
    db_payment = models.PaymentHistory(**payment.dict(), due=due)
    db.add(db_payment)
    db.commit()
    db.refresh(db_payment)
    return db_payment

@app.get("/payments/student/{student_id}", response_model=schemas.StudentPaymentHistory)
def get_student_payment_history(student_id: str, db: Session = Depends(database.get_db)):
    try:
        student = db.query(models.Student).filter(models.Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="Student not found")
        
        payments = db.query(models.PaymentHistory).filter(models.PaymentHistory.student_id == student_id).all()
        pydantic_payments = [schemas.PaymentHistory.from_orm(payment) for payment in payments]
        return schemas.StudentPaymentHistory(student_id=student_id, payments=pydantic_payments)
    except Exception as e:
        logger.error(f"Error in get_student_payment_history: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/payments/year/{year}", response_model=List[schemas.PaymentHistory])
def get_yearly_payments(year: int, db: Session = Depends(database.get_db)):
    try:
        payments = db.query(models.PaymentHistory).filter(extract('year', models.PaymentHistory.date) == year).all()
        if not payments:
            raise HTTPException(status_code=404, detail="No payments found for this year")
        return [schemas.PaymentHistory.from_orm(payment) for payment in payments]
    except Exception as e:
        logger.error(f"Error in get_yearly_payments: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/payments/month/{year}/{month}", response_model=schemas.MonthlyPaymentSummary)
def get_monthly_payments(year: int, month: int, db: Session = Depends(database.get_db)):
    try:
        payments = db.query(models.PaymentHistory).filter(
            extract('year', models.PaymentHistory.date) == year,
            extract('month', models.PaymentHistory.date) == month
        ).all()
        if not payments:
            raise HTTPException(status_code=404, detail="No payments found for this month")
        pydantic_payments = [schemas.PaymentHistory.from_orm(payment) for payment in payments]
        return schemas.MonthlyPaymentSummary(year=year, month=month, payments=pydantic_payments)
    except Exception as e:
        logger.error(f"Error in get_monthly_payments: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/payments/due", response_model=schemas.DuePaymentSummary)
def get_due_payments(db: Session = Depends(database.get_db)):
    try:
        payments = db.query(models.PaymentHistory).filter(models.PaymentHistory.due > 0).all()
        if not payments:
            raise HTTPException(status_code=404, detail="No due payments found")
        
        pydantic_payments = [schemas.PaymentHistory.from_orm(payment) for payment in payments]
        return schemas.DuePaymentSummary(payments=pydantic_payments)
    except Exception as e:
        logger.error(f"Error in get_due_payments: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/monthly_attendance/{student_id}/{year}/{month}", response_model=schemas.MonthlyAttendance)
def get_monthly_attendance(student_id: str, year: int, month: int, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    attendances = db.query(models.Attendance).filter(
        models.Attendance.student_id == student_id,
        models.Attendance.date.between(date(year, month, 1), date(year, month + 1, 1) - timedelta(days=1))
    ).all()

    total_days = len(attendances)
    present_days = sum(1 for a in attendances if a.present)

    return schemas.MonthlyAttendance(
        student_id=student_id,
        month=month,
        year=year,
        total_days=total_days,
        present_days=present_days
    )

@app.get("/monthly_payment/{student_id}/{year}/{month}", response_model=schemas.MonthlyPayment)
def get_monthly_payment(student_id: str, year: int, month: int, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    payments = db.query(models.PaymentHistory).filter(
        models.PaymentHistory.student_id == student_id,
        extract('year', models.PaymentHistory.date) == year,
        extract('month', models.PaymentHistory.date) == month
    ).all()

    total_payment = sum(payment.payment for payment in payments)
    total_paid = sum(payment.paid for payment in payments)
    total_due = sum(payment.due for payment in payments)

    return schemas.MonthlyPayment(
        student_id=student_id,
        month=month,
        year=year,
        total_payment=total_payment,
        total_paid=total_paid,
        total_due=total_due
    )

@app.delete("/students/{student_id}")
def delete_student(student_id: str, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    db.delete(student)
    db.commit()
    return {"message": "Student deleted successfully"}

@app.delete("/attendance/{student_id}/{date}")
def delete_attendance(student_id: str, date: date, db: Session = Depends(database.get_db)):
    attendance = db.query(models.Attendance).filter(
        models.Attendance.student_id == student_id,
        models.Attendance.date == date
    ).first()
    if not attendance:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    db.delete(attendance)
    db.commit()
    return {"message": "Attendance record deleted successfully"}

@app.delete("/payments/{student_id}/{date}")
def delete_payment(student_id: str, date: date, db: Session = Depends(database.get_db)):
    payment = db.query(models.PaymentHistory).filter(
        models.PaymentHistory.student_id == student_id,
        models.PaymentHistory.date == date
    ).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment record not found")
    db.delete(payment)
    db.commit()
    return {"message": "Payment record deleted successfully"}

@app.put("/students/{student_id}", response_model=schemas.Student)
def update_student(student_id: str, student: schemas.StudentCreate, db: Session = Depends(database.get_db)):
    db_student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if db_student is None:
        raise HTTPException(status_code=404, detail="Student not found")
    
    for key, value in student.dict().items():
        setattr(db_student, key, value)
    
    db.commit()
    db.refresh(db_student)
    return db_student

@app.get("/students/{student_id}/yearly_dues", response_model=dict[int, float])
def get_student_yearly_dues(student_id: str, db: Session = Depends(database.get_db)):
    try:
        student = db.query(models.Student).filter(models.Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="Student not found")
        
        yearly_dues = db.query(
            extract('year', models.PaymentHistory.date).label('year'),
            func.sum(models.PaymentHistory.due).label('total_due')
        ).filter(
            models.PaymentHistory.student_id == student_id
        ).group_by(
            extract('year', models.PaymentHistory.date)
        ).all()
        
        dues_dict = {year: float(total_due) for year, total_due in yearly_dues}
        return dues_dict
    except Exception as e:
        logger.error(f"Error in get_student_yearly_dues: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.post("/exams/", response_model=schemas.ExamHistory)
def create_exam(exam: schemas.ExamHistoryCreate, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == exam.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    db_exam = models.ExamHistory(**exam.dict())
    db.add(db_exam)
    db.commit()
    db.refresh(db_exam)
    return db_exam

@app.get("/exams/student/{student_id}", response_model=schemas.StudentExamHistory)
def get_student_exam_history(student_id: str, db: Session = Depends(database.get_db)):
    try:
        student = db.query(models.Student).filter(models.Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="Student not found")
        
        exams = db.query(models.ExamHistory).filter(models.ExamHistory.student_id == student_id).all()
        return schemas.StudentExamHistory(student_id=student_id, exams=exams)
    except Exception as e:
        logger.error(f"Error in get_student_exam_history: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/exams/year/{year}", response_model=schemas.YearlyExamSummary)
def get_yearly_exams(year: int, db: Session = Depends(database.get_db)):
    exams = db.query(models.ExamHistory).filter(extract('year', models.ExamHistory.date) == year).all()
    if not exams:
        raise HTTPException(status_code=404, detail="No exams found for this year")
    return schemas.YearlyExamSummary(year=year, exams=exams)

@app.get("/exams/month/{year}/{month}", response_model=schemas.MonthlyExamSummary)
def get_monthly_exams(year: int, month: int, db: Session = Depends(database.get_db)):
    exams = db.query(models.ExamHistory).filter(
        extract('year', models.ExamHistory.date) == year,
        extract('month', models.ExamHistory.date) == month
    ).all()
    if not exams:
        raise HTTPException(status_code=404, detail="No exams found for this month")
    return schemas.MonthlyExamSummary(year=year, month=month, exams=exams)

@app.get("/exams/percentage/{student_id}/{year}/{month}", response_model=schemas.MonthlyExamPercentage)
def get_monthly_exam_percentage(student_id: str, year: int, month: int, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    exams = db.query(models.ExamHistory).filter(
        models.ExamHistory.student_id == student_id,
        extract('year', models.ExamHistory.date) == year,
        extract('month', models.ExamHistory.date) == month
    ).all()

    if not exams:
        raise HTTPException(status_code=404, detail="No exams found for this student in the specified month")

    total_marks = sum(exam.total_marks for exam in exams)
    obtained_marks = sum(exam.obtained_marks for exam in exams)
    percentage = (obtained_marks / total_marks) * 100 if total_marks > 0 else 0

    return schemas.MonthlyExamPercentage(
        student_id=student_id,
        year=year,
        month=month,
        percentage=round(percentage, 2)
    )

@app.get("/students/{student_id}/payment_history_excel")
def get_student_payment_history_excel(student_id: str, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    payments = db.query(models.PaymentHistory).filter(models.PaymentHistory.student_id == student_id).all()

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Payment History"

    # Add student information
    sheet.merge_cells('A1:E1')
    sheet['A1'] = f"Payment History for {student.name} (ID: {student.id})"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Add headers
    headers = ["Date", "Payment", "Paid", "Due", "Total Subjects"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add payment data and calculate totals
    total_payment = 0
    total_paid = 0
    total_due = 0
    for row, payment in enumerate(payments, start=4):
        sheet.cell(row=row, column=1, value=payment.date).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=2, value=payment.payment).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=3, value=payment.paid).alignment = Alignment(horizontal='right')
        due_cell = sheet.cell(row=row, column=4, value=payment.due)
        due_cell.alignment = Alignment(horizontal='right')
        if payment.due > 0:
            due_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            due_cell.font = Font(color="9C0006")
        else:
            due_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            due_cell.font = Font(color="006100")
        sheet.cell(row=row, column=5, value=payment.total_subjects).alignment = Alignment(horizontal='center')
        
        total_payment += payment.payment
        total_paid += payment.paid
        total_due += payment.due

    # Apply borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet['A3:E' + str(sheet.max_row)]:
        for cell in row:
            cell.border = thin_border

    # Adjust column widths
    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 15

    # Add total row
    total_row = sheet.max_row + 2
    sheet.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    sheet.cell(row=total_row, column=2, value=total_payment).font = Font(bold=True)
    sheet.cell(row=total_row, column=3, value=total_paid).font = Font(bold=True)
    sheet.cell(row=total_row, column=4, value=total_due).font = Font(bold=True)

    # Create a temporary file to save the Excel sheet
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    # Return the file as a downloadable response
    return FileResponse(
        tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"payment_history_{student_id}.xlsx"
    )

@app.get("/students/{student_id}/exam_history_excel")
def get_student_exam_history_excel(student_id: str, db: Session = Depends(database.get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    exams = db.query(models.ExamHistory).filter(models.ExamHistory.student_id == student_id).order_by(models.ExamHistory.date).all()

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Exam History"

    # Add student information
    sheet.merge_cells('A1:E1')
    sheet['A1'] = f"Exam History for {student.name} (ID: {student.id})"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Add headers
    headers = ["Date", "Subject", "Total Marks", "Obtained Marks", "Percentage"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add exam data
    for row, exam in enumerate(exams, start=4):
        sheet.cell(row=row, column=1, value=exam.date).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=2, value=exam.subject_name).alignment = Alignment(horizontal='left')
        sheet.cell(row=row, column=3, value=exam.total_marks).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=4, value=exam.obtained_marks).alignment = Alignment(horizontal='right')
        percentage = (exam.obtained_marks / exam.total_marks) * 100 if exam.total_marks > 0 else 0
        sheet.cell(row=row, column=5, value=round(percentage, 2)).alignment = Alignment(horizontal='right')

    # Apply borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet['A3:E' + str(sheet.max_row)]:
        for cell in row:
            cell.border = thin_border

    # Adjust column widths
    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 15

    # Create a line chart
    chart = LineChart()
    chart.title = "Exam Performance Over Time"
    chart.y_axis.title = "Percentage"
    chart.x_axis.title = "Exams"

    # Add data to the chart
    data = Reference(sheet, min_col=5, min_row=3, max_row=sheet.max_row, max_col=5)
    chart.add_data(data, titles_from_data=True)

    # Set categories (x-axis labels) to exam dates
    dates = Reference(sheet, min_col=1, min_row=4, max_row=sheet.max_row)
    chart.set_categories(dates)

    # Add the chart to the sheet
    sheet.add_chart(chart, "G3")

    # Create a temporary file to save the Excel sheet
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    # Return the file as a downloadable response
    return FileResponse(
        tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"exam_history_{student_id}.xlsx"
    )